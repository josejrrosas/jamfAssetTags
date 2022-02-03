#!/usr/bin/env python3
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as wait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import time
from keyboard import press
import openpyxl
from openpyxl import load_workbook
import string
import tkinter as tk
from tkinter import *
import sys

class MyWindow:
    def __init__(self, win):
        #Labels
        single = r'Enter Excel Sheet Path (Ex:C:\Users\name\Desktop\file.xlsx)'
        double = single.replace('\\', '\\\\')
        self.excelLabel=Label(win, text= double)
        self.userLabel=Label(win, text='Email/Username')
        self.passLabel=Label(win, text='Password')
        self.firstLabel=Label(win, text='Enter first row number')
        self.lastLabel=Label(win, text='Enter last row number')
        self.columnLabel=Label(win, text='Enter Serial Column Number')
        self.assetLabel=Label(win, text='Enter Asset Column Number')
        
        #Entrys
        self.excelEntry=Entry()
        self.userEntry=Entry()
        self.passEntry=Entry(show="*")
        self.firstEntry=Entry()
        self.lastEntry=Entry()
        self.columnEntry=Entry()
        self.assetEntry=Entry()

        #Placement
        self.excelLabel.place(x=100, y=50)
        self.excelEntry.place(x=100, y=70)
        
        self.userLabel.place(x=100, y=100)
        self.userEntry.place(x=100, y=120)
        
        self.passLabel.place(x=100, y=150)
        self.passEntry.place(x=100, y=170)

        self.firstLabel.place(x=100, y=200)
        self.firstEntry.place(x=100, y=220)

        self.lastLabel.place(x=100, y=250)
        self.lastEntry.place(x=100, y=270)

        self.columnLabel.place(x=100, y=300)
        self.columnEntry.place(x=100, y=320)

        self.assetLabel.place(x=100, y=350)
        self.assetEntry.place(x=100, y=370)

        #Buttons
        self.btn1 = Button(win, text='Submit')
        self.b1=Button(win, text='Submit',command=self.change)
        self.b1.place(x=100, y=450)

    def change(self):
        #retrieve entered info
        username=self.userEntry.get()
        password=self.passEntry.get()
        excelPath=self.excelEntry.get()
       #get cell string and parse to int
        serialColumnStr=self.columnEntry.get()
        assetColumnStr=self.assetEntry.get()
        serialColumn=int(serialColumnStr)
        assetColumn=int(assetColumnStr)
        #get cell string and parse to int
        firstCellStr=self.firstEntry.get()
        lastCellStr=self.lastEntry.get()
        firstCell=int(firstCellStr)
        lastCell=int(lastCellStr)+1

        path = excelPath
        driver = webdriver.Chrome(ChromeDriverManager().install())
        driver.get('https://americanairlines2.jamfcloud.com/')
        driver.maximize_window()

        userForm = '//*[@id="userNameInput"]'
        nextButton = '//*[@id="idSIButton9"]'
        emailForm = '//*[@id="i0116"]'
        passForm = '//*[@id="passwordInput"]'
        loginButton = '//*[@id="submitButton"]'
        noBtn = '//*[@id="idBtn_Back"]'
        devicesBtn = '//*[@id="side-nav-tabs"]/li[2]/a'
        searchBtn = '//*[@id="query"]'
        userName = username
        passWord = password

        #click email form and enter email and click next
        time.sleep(3)
        driver.find_element_by_xpath(emailForm).click()
        driver.find_element_by_xpath(emailForm).send_keys(userName)
        driver.find_element_by_xpath(nextButton).click()

        #click password form and enter password and click login button
        time.sleep(5)
        driver.find_element_by_xpath(passForm).click()
        driver.find_element_by_xpath(passForm).send_keys(passWord)
        driver.find_element_by_xpath(loginButton).click()

        #click no button
        time.sleep(1.5)
        driver.find_element_by_xpath(noBtn).click()

       #click devices button
        time.sleep(7)
        driver.find_element_by_xpath(devicesBtn).click()    

        #loop through serial and asset in excel
        wb_obj = openpyxl.load_workbook(path, data_only=True)
        sheet_obj = wb_obj.active
        
        for row in range(firstCell, lastCell):
            serialCell = sheet_obj.cell(row,column=serialColumn).value
            assetCell = sheet_obj.cell(row,column=assetColumn).value
            print(row, " " ,  serialCell ,' ', assetCell)

            #>>>>>>>>>>>>>>>>>LOOP STARTS HERE<<<<<<<<<<<<<<<<<<<<<<<<<
            
            #click search form and enter serial 
            time.sleep(3)
            wait(driver, 5).until(EC.frame_to_be_available_and_switch_to_it((By.XPATH,'//*[@id="main-ui-view"]/div/div/div/jamf-legacy-view/iframe')))
            wait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, searchBtn))).send_keys(serialCell)
            press('enter')

            #click device name to view device info 
            wait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="mobileDevicesTable"]/tbody/tr/td/a'))).click()

            #click edit button to edit device info
            driver.switch_to.default_content()
            iframe = driver.find_element_by_xpath('//*[@id="main-ui-view"]/div/div/div/jamf-legacy-view/iframe')
            driver.switch_to.frame(iframe)
            wait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="General_Pane"]/div/div/div[1]/button'))).click()
            time.sleep(1)

            #clear and fill asset tag field
            wait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="FIELD_ASSET_TAG"]'))).clear()
            time.sleep(1)
            wait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="FIELD_ASSET_TAG"]'))).send_keys(assetCell)

            #save asset tag
            time.sleep(1)
            wait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="Edit_General_Pane"]/div/div/div[1]/input[2]'))).click()

            #click devices button again
            time.sleep(1)
            driver.switch_to.default_content()
            wait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="side-nav-tabs"]/li[2]/a/img[2]'))).click()
            #>>>>>>>>>>>>>>>>>>>>LOOP RESTARTS OR ENDS HERE <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

window=Tk()
mywin=MyWindow(window)
window.title('JAMF Asset Tag')
window.geometry("500x500+10+10")
window.mainloop()
print('done')

input('Press ENTER to exit')


